#!/usr/bin/env python
#-*-conding utf-8-*-

from enum import Enum
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os, sys
import time

FilesToCheck = [".h", ".hpp", ".c", ".cpp"]
class Status(Enum):
    Init = 0
    Common = 1
    CharString = 2
    PreComment = 3
    LineComment = 4
    BlockComments = 5
    PreExitComment = 6
    PreCombination = 7
    Combination = 8


def CuntFileCodeLine(fileName):
    if not os.path.isfile(fileName):
        print "%s is not a file name" % fileName
        return
    currentStatus = Status.Common
    preCombinationLastStatus = Status.Common
    lineNum = 0
    linecunt = 0
    with open(fileName) as obFile:
        for line in obFile:
            linecunt += 1
            # print linecunt,lineNum
            isNewLine = True
            lineCode = line.strip()
            if len(lineCode) < 1:
                continue
            lineCode += '\n'
            for obChar in lineCode:
                if currentStatus == Status.Common:
                    if r"/" == obChar:
                        currentStatus = Status.PreComment
                    elif r'"' == obChar:
                        currentStatus = Status.CharString
                    elif "\n" == obChar:
                        pass
                    else:
                        if isNewLine == True:
                            lineNum += 1
                            isNewLine = False
                elif currentStatus == Status.PreComment:
                    if r"/" == obChar:
                        currentStatus = Status.LineComment
                    elif r"*" == obChar:
                        currentStatus = Status.BlockComments
                    else:
                        currentStatus = Status.Common         # back to common status
                elif currentStatus == Status.BlockComments:
                    if r"*" == obChar:
                        currentStatus = Status.PreExitComment
                    else:
                        pass
                elif currentStatus == Status.LineComment:
                    if "\n" == obChar:
                        currentStatus = Status.Common
                    elif "\\" == obChar:
                        preCombinationLastStatus = currentStatus
                        currentStatus = Status.PreCombination
                    else:
                        pass
                elif currentStatus == Status.CharString:
                    if r'"' == obChar:
                        currentStatus = Status.Common
                    elif "\\" == obChar:
                        preCombinationLastStatus = currentStatus
                        currentStatus = Status.PreCombination
                    else:
                        pass
                elif currentStatus == Status.PreExitComment:
                    if "/" == obChar:
                        currentStatus = Status.Common
                    else:
                        currentStatus = Status.BlockComments
                elif currentStatus == Status.PreCombination:
                    currentStatus = preCombinationLastStatus

    return lineNum

def CuntDirCodeLine(DirName):
    totalCnt = 0
    fileCnt = 0
    if not os.path.isdir(DirName):
        print "%s is not a dir name !" % DirName
    for root,dirs,files in os.walk(DirName):
        for file in files:
            if os.path.splitext(file)[1] in FilesToCheck:
                file = os.path.join(root, file)
                fileLineCnt = CuntFileCodeLine(file)
                if None != fileLineCnt:
                    print "%s: %d" % (file, fileLineCnt)
                    totalCnt += fileLineCnt
                    fileCnt += 1
                    ws.cell(column = 1, row = fileCnt + 1, value = file)
                    ws.cell(column = 2, row = fileCnt + 1, value = fileLineCnt)
    print fileCnt, totalCnt
    ws.cell(column=1, row=fileCnt + 3, value = "total file num: %d" % fileCnt)
    ws.cell(column=2, row=fileCnt + 3, value = "total line num: %d" % totalCnt)

def mainFunction():
    dirname = raw_input("Pleas input the target dir: ")
    execelName = raw_input("Please input the output file name (.xlsx): ")
    if len(execelName) > 0:
        execelName = execelName.rstrip(".xlsx") + ".xlsx"
    else:
        execelName = "default.xlsx"
    print execelName
    wb = Workbook()
    global ws
    ws = wb.create_sheet("LineCount")
    ws["A1"] = "file name"
    ws["B1"] = "line count"
    if len(dirname) > 0:
        CuntDirCodeLine(dirname)
    else:
        CuntDirCodeLine(".")
    ws.column_dimensions[get_column_letter(1)].width = 33
    ws.column_dimensions[get_column_letter(2)].width = 19
    wb.save(filename=execelName)

if __name__ == "__main__":
    start = time.time()
    # mainFunction()
    tt = CuntFileCodeLine("/home/shigm/cc++/aa.cpp")
    print tt
    dur = time.time() - start
    print "Cost %d seconds !" % dur